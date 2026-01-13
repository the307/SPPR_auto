from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import pandas as pd


def export_to_excel(
    result_df: pd.DataFrame,
    output_path: str,
    calc_date,
    alarm_flag: bool = False,
    alarm_msg: str | None = None,
    month_column_name: str = "F_bp_month"
):
    """
    Сохраняет DataFrame в Excel и при необходимости:
    - исключает служебные колонки (*_data)
    - красит ячейку месячного значения
    - добавляет комментарий

    alarm_flag / alarm_msg — служебные сигналы из calculate
    """

    # =========================================================
    # 1. Исключаем служебные колонки (_data)
    # (все колонки экспортируются, служебные колонки можно добавить фильтрацию при необходимости)
    export_df = result_df.copy()

    # =========================================================
    # 2. Сохраняем DataFrame
    export_df.to_excel(output_path, index=False)

    # Если тревоги нет — на этом всё
    if not alarm_flag:
        print(f"Результат сохранён в {output_path}")
        return

    # =========================================================
    # 3. Открываем Excel для форматирования
    wb = load_workbook(output_path)
    ws = wb.active

    # =========================================================
    # 4. Находим колонку месячного значения
    headers = [cell.value for cell in ws[1]]
    if month_column_name not in headers:
        wb.save(output_path)
        print(
            f"Колонка '{month_column_name}' не найдена — файл сохранён без подсветки"
        )
        return

    col_idx = headers.index(month_column_name) + 1

    # =========================================================
    # 5. Ищем строку расчётной даты
    target_row = None
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value == calc_date:
            target_row = row
            break

    if target_row is None:
        wb.save(output_path)
        print("Дата расчёта не найдена — файл сохранён без подсветки")
        return

    # =========================================================
    # 6. Красим и добавляем комментарий
    cell = ws.cell(row=target_row, column=col_idx)
    cell.fill = PatternFill("solid", fgColor="FF9999")
    cell.comment = Comment(alarm_msg or "Контрольное условие не выполнено", "СППР")

    wb.save(output_path)
    print(f"Результат сохранён в {output_path} (с предупреждением)")