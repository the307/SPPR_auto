#Функция превода данных из JSON в таблцу Excel

import json
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Чтение json
json_path = Path("data_output.json")
if not json_path.exists():
    json_path = Path("output.json")
with json_path.open("r", encoding="utf-8") as f:
    dist_json = json.load(f)

error = dist_json.get("error")
if error:
    raise SystemExit(f'Ошибка расчета: {error.get("message")}')

control = dist_json.get("control") or {}
month = dist_json.get("month") or {}
last_day = dist_json.get("last_day") or {}
days = dist_json.get("days") or []


# Чтение excel шаблона
excel_file = load_workbook('balance_sample.xlsx')
excel_sheet = excel_file.active


# Создаем стили заливки
# Вычисляемые значения
pale_green_fill = PatternFill(start_color='DAEEAD', end_color='DAEEAD', fill_type='solid')
# Вводимые значения
pale_blue_fill = PatternFill(start_color='BDE8F5', end_color='BDE8F5', fill_type='solid')
# Проверка не пройдена
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
# Проверка пройдена
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')


def _cell_value(value):
    if isinstance(value, dict):
        if "value" in value:
            return _cell_value(value.get("value"))
        return None
    return value


def _status(value):
    if isinstance(value, dict):
        return value.get("status", 0)
    return 0


def _format_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, str):
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt).strftime("%d.%m.%Y")
            except ValueError:
                continue
    return value


def _parse_day_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None

def _last(key, alt_key=None):
    value = last_day.get(key)
    if value is None and alt_key:
        value = last_day.get(alt_key)
    return _cell_value(value)

def _m(key):
    return _cell_value(month.get(key))

# Даты для Excel и выборка предыдущего дня
days_sorted = sorted(
    days,
    key=lambda item: _parse_day_date(item.get("date")) or datetime.max,
)
control_date = _parse_day_date(control.get("Date_calculation"))
prev_day_json = next(
    (item for item in days_sorted if _parse_day_date(item.get("date")) == control_date),
    None,
)

calc_date = None
if days_sorted:
    last_day_date = days_sorted[-1].get("date")
    if last_day_date:
        try:
            calc_date = datetime.strptime(last_day_date, "%Y-%m-%d")
        except ValueError:
            calc_date = None
if calc_date is None:
    calc_date = datetime.today()
date_str = control.get("Date_calculation") or calc_date.strftime("%Y-%m-%d")
start_autobalance = bool(control.get("Start_autobalance"))

# Записываем в Excel значения на начало месяца
row = 4

excel_sheet['A'+ str(row)] = _format_date(control.get("Date_calculation"))
excel_sheet['A'+ str(row)].fill = pale_blue_fill

excel_sheet['F'+ str(row)] = _last("V_upsv_yu_0")
excel_sheet['F'+ str(row)].fill = pale_blue_fill

excel_sheet['G'+ str(row)] = _last("V_upsv_s_0")
excel_sheet['G'+ str(row)].fill = pale_blue_fill

excel_sheet['H'+ str(row)] = _last("V_cps_0")
excel_sheet['H'+ str(row)].fill = pale_blue_fill

V_cppn_1_0 = _last("V_cppn_1_0")
if V_cppn_1_0 is None and prev_day_json is not None:
    V_cppn_1_0 = _cell_value(prev_day_json.get("V_cppn_1"))
excel_sheet['I'+ str(row)] = V_cppn_1_0
excel_sheet['I'+ str(row)].fill = pale_green_fill

excel_sheet['J'+ str(row)] = _last("V_lodochny_cps_upsv_0", "V_lodochny_cps_upsv_yu_prev")
excel_sheet['J'+ str(row)].fill = pale_blue_fill

excel_sheet['V'+ str(row)] = _last("V_upn_suzun_0")
excel_sheet['V'+ str(row)].fill = pale_blue_fill

v_suzun_slu_0 = _last("V_suzun_slu_0")
if v_suzun_slu_0 is None and prev_day_json is not None:
    v_suzun_slu_0 = _cell_value(prev_day_json.get("V_suzun_slu"))
excel_sheet['W'+ str(row)] = v_suzun_slu_0
excel_sheet['W'+ str(row)].fill = pale_green_fill

excel_sheet['X'+ str(row)] = _last("V_suzun_vslu_0")
excel_sheet['X'+ str(row)].fill = pale_blue_fill

excel_sheet['Y'+ str(row)] = _last("V_suzun_tng_0")
excel_sheet['Y'+ str(row)].fill = pale_blue_fill

excel_sheet['AL'+ str(row)] = _last("V_tagul_tr_0", "V_tagul_tp_0")
excel_sheet['AL'+ str(row)].fill = pale_blue_fill

excel_sheet['AR'+ str(row)] = _last("V_upn_lodochny_0")
excel_sheet['AR'+ str(row)].fill = pale_blue_fill

excel_sheet['AS'+ str(row)] = _last("V_lodochny_0")
excel_sheet['AS'+ str(row)].fill = pale_blue_fill

excel_sheet['AT'+ str(row)] = _last("V_ichem_0")
excel_sheet['AT'+ str(row)].fill = pale_blue_fill

excel_sheet['BF'+ str(row)] = _last("V_gnps_0")
excel_sheet['BF'+ str(row)].fill = pale_blue_fill

excel_sheet['BG'+ str(row)] = _last("V_nps_1_0")
excel_sheet['BG'+ str(row)].fill = pale_blue_fill

excel_sheet['BH'+ str(row)] = _last("V_nps_2_0")
excel_sheet['BH'+ str(row)].fill = pale_blue_fill

excel_sheet['BI'+ str(row)] = _last("V_knps_0", "V_knsp_0")
excel_sheet['BI'+ str(row)].fill = pale_blue_fill

excel_sheet['BJ'+ str(row)] = _last("V_tstn_0")
excel_sheet['BJ'+ str(row)].fill = pale_green_fill

excel_sheet['BK'+ str(row)] = _last("V_tstn_rn_vn_0")
excel_sheet['BK'+ str(row)].fill = pale_blue_fill

excel_sheet['BL'+ str(row)] = _last("V_tstn_vn_0")
excel_sheet['BL'+ str(row)].fill = pale_green_fill

excel_sheet['BM'+ str(row)] = _last("V_tstn_suzun_0")
excel_sheet['BM'+ str(row)].fill = pale_green_fill

excel_sheet['BN'+ str(row)] = _last("V_tstn_suzun_vankor_0")
excel_sheet['BN'+ str(row)].fill = pale_blue_fill

excel_sheet['BO'+ str(row)] = _last("V_tstn_suzun_vslu_0")
excel_sheet['BO'+ str(row)].fill = pale_blue_fill

excel_sheet['BP'+ str(row)] = _last("V_tstn_tagul_obsh_0")
excel_sheet['BP'+ str(row)].fill = pale_green_fill

excel_sheet['BQ'+ str(row)] = _last("V_tstn_lodochny_0")
excel_sheet['BQ'+ str(row)].number_format = "0.000"
excel_sheet['BQ'+ str(row)].fill = pale_blue_fill

excel_sheet['BR'+ str(row)] = _last("V_tstn_tagul_0")
excel_sheet['BR'+ str(row)].fill = pale_blue_fill

excel_sheet['BS'+ str(row)] = _last("V_tstn_skn_0")
excel_sheet['BS'+ str(row)].fill = pale_blue_fill

excel_sheet['BT'+ str(row)] = _last("V_tstn_vo_0")
excel_sheet['BT'+ str(row)].fill = pale_blue_fill

excel_sheet['BU'+ str(row)] = _last("V_tstn_tng_0")
excel_sheet['BU'+ str(row)].fill = pale_blue_fill

excel_sheet['BV'+ str(row)] = _last("V_tstn_kchng_0")
excel_sheet['BV'+ str(row)].fill = pale_blue_fill

row += 1

# Записываем в Excel значения по дням месяца
for day_json in days_sorted:
    def _d(key):
        return _cell_value(day_json.get(key))

    excel_sheet['A'+ str(row)] = _format_date(_d("date"))
    excel_sheet['A'+ str(row)].fill = pale_blue_fill
    
    excel_sheet['D'+ str(row)] = _d("Q_vankor")
    excel_sheet['D'+ str(row)].fill = pale_blue_fill
    
    excel_sheet['F'+ str(row)] = _cell_value(day_json.get("V_upsv_yu"))
    match _status(day_json.get("V_upsv_yu")):
        case 0:
            excel_sheet['F'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['F'+ str(row)].fill = green_fill
        case _:
            excel_sheet['F'+ str(row)].fill = red_fill
    
    excel_sheet['G'+ str(row)] = _cell_value(day_json.get("V_upsv_s"))
    match _status(day_json.get("V_upsv_s")):
        case 0:
            excel_sheet['G'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['G'+ str(row)].fill = green_fill
        case _:
            excel_sheet['G'+ str(row)].fill = red_fill

    excel_sheet['H'+ str(row)] = _cell_value(day_json.get("V_cps"))
    match _status(day_json.get("V_cps")):
        case 0:
            excel_sheet['H'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['H'+ str(row)].fill = green_fill
        case _:
            excel_sheet['H'+ str(row)].fill = red_fill

    excel_sheet['I'+ str(row)] = _d("V_cppn_1")
    excel_sheet['I'+ str(row)].fill = pale_green_fill

    excel_sheet['J'+ str(row)] = _cell_value(day_json.get("V_lodochny_cps_upsv_yu"))
    match _status(day_json.get("V_lodochny_cps_upsv_yu")):
        case 0:
            excel_sheet['J'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['J'+ str(row)].fill = green_fill
        case _:
            excel_sheet['J'+ str(row)].fill = red_fill

    excel_sheet['K'+ str(row)] = _d("G_payaha")
    excel_sheet['K'+ str(row)].fill = pale_blue_fill

    excel_sheet['L'+ str(row)] = _d("G_sikn")
    excel_sheet['L'+ str(row)].fill = pale_green_fill
    
    excel_sheet['M'+ str(row)] = _d("G_sikn_vankor")
    excel_sheet['M'+ str(row)].fill = pale_green_fill
    
    excel_sheet['N'+ str(row)] = _d("G_sikn_suzun")
    excel_sheet['N'+ str(row)].fill = pale_green_fill
    
    excel_sheet['O'+ str(row)] = _d("G_sikn_vslu")
    excel_sheet['O'+ str(row)].fill = pale_green_fill
    
    excel_sheet['P'+ str(row)] = _cell_value(day_json.get("G_sikn_tagul"))
    match _status(day_json.get("G_sikn_tagul")):
        case 0:
            excel_sheet['P'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['P'+ str(row)].fill = green_fill
        case _:
            excel_sheet['P'+ str(row)].fill = red_fill
    
    excel_sheet['Q'+ str(row)] = _d("G_sikn_tng")
    excel_sheet['Q'+ str(row)].fill = pale_green_fill

    excel_sheet['R'+ str(row)] = _d("delta_G_sikn")
    excel_sheet['R'+ str(row)].fill = pale_green_fill

    excel_sheet['T'+ str(row)] = _d("Q_suzun")
    excel_sheet['T'+ str(row)].fill = pale_blue_fill

    excel_sheet['U'+ str(row)] = _d("Q_vslu")
    excel_sheet['U'+ str(row)].fill = pale_blue_fill

    excel_sheet['V'+ str(row)] = _cell_value(day_json.get("V_upn_suzun"))
    match _status(day_json.get("V_upn_suzun")):
        case 0:
            excel_sheet['V'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['V'+ str(row)].fill = green_fill
        case _:
            excel_sheet['V'+ str(row)].fill = red_fill


    excel_sheet['W'+ str(row)] = _d("V_suzun_slu")
    excel_sheet['W'+ str(row)].fill = pale_green_fill

    excel_sheet['X'+ str(row)] = _d("V_suzun_vslu")
    excel_sheet['X'+ str(row)].fill = pale_green_fill

    excel_sheet['Y'+ str(row)] = _d("V_suzun_tng")
    excel_sheet['Y'+ str(row)].fill = pale_green_fill

    excel_sheet['Z'+ str(row)] = _d("G_suzun")
    excel_sheet['Z'+ str(row)].fill = pale_green_fill

    excel_sheet['AA'+ str(row)] = _d("G_suzun_slu")
    excel_sheet['AA'+ str(row)].fill = pale_green_fill
    
    excel_sheet['AB'+ str(row)] = _d("G_suzun_vslu")
    excel_sheet['AB'+ str(row)].fill = pale_green_fill
    
    excel_sheet['AC'+ str(row)] = _d("G_suzun_tng")
    excel_sheet['AC'+ str(row)].fill = pale_blue_fill
    
    excel_sheet['AD'+ str(row)] = _d("delta_G_suzun")
    excel_sheet['AD'+ str(row)].fill = pale_green_fill
    
    excel_sheet['AE'+ str(row)] = _d("Q_tng")
    excel_sheet['AE'+ str(row)].fill = pale_blue_fill
    
    excel_sheet['AF'+ str(row)] = _d("G_payaha")
    excel_sheet['AF'+ str(row)].fill = pale_blue_fill
    
    excel_sheet['AG'+ str(row)] = _d("G_buying_oil")
    excel_sheet['AG'+ str(row)].fill = pale_green_fill
    
    excel_sheet['AH'+ str(row)] = _d("G_out_udt")
    excel_sheet['AH'+ str(row)].fill = pale_green_fill
        
    excel_sheet['AI'+ str(row)] = _d("G_per")
    excel_sheet['AI'+ str(row)].fill = pale_green_fill

    excel_sheet['AK'+ str(row)] = _d("Q_tagul")
    excel_sheet['AK'+ str(row)].fill = pale_blue_fill
    
    excel_sheet['AL'+ str(row)] = _cell_value(day_json.get("V_tagul_tr"))
    match _status(day_json.get("V_tagul_tr")):
        case 0:
            excel_sheet['AL'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['AL'+ str(row)].fill = green_fill
        case _:
            excel_sheet['AL'+ str(row)].fill = red_fill
    
    excel_sheet['AM'+ str(row)] = _d("G_tagul")
    excel_sheet['AM'+ str(row)].fill = pale_green_fill
    
    excel_sheet['AN'+ str(row)] = _d("delta_G_tagul")
    excel_sheet['AN'+ str(row)].fill = pale_green_fill

    excel_sheet['AP'+ str(row)] = _d("Q_lodochny")
    excel_sheet['AP'+ str(row)].fill = pale_blue_fill
    
    g_lodochny_upsv_yu = _d("G_lodochny_upsv_yu")
    if g_lodochny_upsv_yu is None:
        g_lodochny_upsv_yu = _d("G_lodochny_uspv_yu")
    excel_sheet['AQ'+ str(row)] = g_lodochny_upsv_yu
    excel_sheet['AQ'+ str(row)].fill = pale_green_fill
        
    excel_sheet['AR'+ str(row)] = _cell_value(day_json.get("V_upn_lodochny"))
    match _status(day_json.get("V_upn_lodochny")):
        case 0:
            excel_sheet['AR'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['AR'+ str(row)].fill = green_fill
        case _:
            excel_sheet['AR'+ str(row)].fill = red_fill

    excel_sheet['AS'+ str(row)] = _d("V_lodochny")
    excel_sheet['AS'+ str(row)].fill = pale_green_fill
    
    excel_sheet['AT'+ str(row)] = _cell_value(day_json.get("V_ichem"))
    match _status(day_json.get("V_ichem")):
        case 0:
            excel_sheet['AT'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['AT'+ str(row)].fill = green_fill
        case _:
            excel_sheet['AT'+ str(row)].fill = red_fill

    
    excel_sheet['AU'+ str(row)] = _d("G_upn_lodochny")
    excel_sheet['AU'+ str(row)].fill = pale_green_fill
    
    excel_sheet['AV'+ str(row)] = _d("G_lodochny")
    excel_sheet['AV'+ str(row)].fill = pale_green_fill
    
    excel_sheet['AW'+ str(row)] = _d("G_ichem")
    excel_sheet['AW'+ str(row)].fill = pale_blue_fill

    excel_sheet['AX'+ str(row)] = _d("delta_G_upn_lodochny")
    excel_sheet['AX'+ str(row)].fill = pale_green_fill
    
    excel_sheet['AY'+ str(row)] = _d("Q_vo")
    excel_sheet['AY'+ str(row)].fill = pale_blue_fill
    
    excel_sheet['AZ'+ str(row)] = _d("G_upn_lodochny_ichem")
    excel_sheet['AZ'+ str(row)].fill = pale_green_fill
    
    excel_sheet['BA'+ str(row)] = _d("G_tagul_lodochny")
    excel_sheet['BA'+ str(row)].fill = pale_green_fill
    
    excel_sheet['BB'+ str(row)] = _d("Q_kchng")
    excel_sheet['BB'+ str(row)].fill = pale_blue_fill
    
    excel_sheet['BC'+ str(row)] = _d("G_kchng")
    excel_sheet['BC'+ str(row)].fill = pale_green_fill
    
    excel_sheet['BD'+ str(row)] = _d("G_skn")
    excel_sheet['BD'+ str(row)].fill = pale_blue_fill
    
    excel_sheet['BE'+ str(row)] = _d("G_gnps")
    excel_sheet['BE'+ str(row)].fill = pale_green_fill
    
    excel_sheet['BF'+ str(row)] = _cell_value(day_json.get("V_gnps"))
    match _status(day_json.get("V_gnps")):
        case 0:
            excel_sheet['BF'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BF'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BF'+ str(row)].fill = red_fill
    
    excel_sheet['BG'+ str(row)] = _cell_value(day_json.get("V_nps_1"))
    match _status(day_json.get("V_nps_1")):
        case 0:
            excel_sheet['BG'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BG'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BG'+ str(row)].fill = red_fill
    
    excel_sheet['BH'+ str(row)] = _cell_value(day_json.get("V_nps_2"))
    match _status(day_json.get("V_nps_2")):
        case 0:
            excel_sheet['BH'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BH'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BH'+ str(row)].fill = red_fill
    
    v_knps = day_json.get("V_knps")
    if v_knps is None:
        v_knps = day_json.get("V_knsp")
    excel_sheet['BI'+ str(row)] = _cell_value(v_knps)
    match _status(v_knps):
        case 0:
            excel_sheet['BI'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BI'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BI'+ str(row)].fill = red_fill
    
    excel_sheet['BJ'+ str(row)] = _d("V_tstn")
    excel_sheet['BJ'+ str(row)].fill = pale_green_fill
    
    excel_sheet['BK'+ str(row)] = _d("V_tstn_rn_vn")
    excel_sheet['BK'+ str(row)].number_format = "0.000"
    excel_sheet['BK'+ str(row)].fill = pale_green_fill
    
    excel_sheet['BL'+ str(row)] = _cell_value(day_json.get("V_tstn_vn"))
    match _status(day_json.get("V_tstn_vn")):
        case 0:
            excel_sheet['BL'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BL'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BL'+ str(row)].fill = red_fill
    
    excel_sheet['BM'+ str(row)] = _cell_value(day_json.get("V_tstn_suzun"))
    excel_sheet['BM'+ str(row)].number_format = "0.000"
    match _status(day_json.get("V_tstn_suzun")):
        case 0:
            excel_sheet['BM'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BM'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BM'+ str(row)].fill = red_fill
    
    excel_sheet['BN'+ str(row)] = _cell_value(day_json.get("V_tstn_suzun_vankor"))
    excel_sheet['BN'+ str(row)].number_format = "0.000"
    match _status(day_json.get("V_tstn_suzun_vankor")):
        case 0:
            excel_sheet['BN'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BN'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BN'+ str(row)].fill = red_fill
    
    excel_sheet['BO'+ str(row)] = _cell_value(day_json.get("V_tstn_suzun_vslu"))
    excel_sheet['BO'+ str(row)].number_format = "0.000"
    match _status(day_json.get("V_tstn_suzun_vslu")):
        case 0:
            excel_sheet['BO'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BO'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BO'+ str(row)].fill = red_fill
    
    excel_sheet['BP'+ str(row)] = _cell_value(day_json.get("V_tstn_tagul_obsh"))
    excel_sheet['BP'+ str(row)].number_format = "0.000"
    match _status(day_json.get("V_tstn_tagul_obsh")):
        case 0:
            excel_sheet['BP'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BP'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BP'+ str(row)].fill = red_fill

    excel_sheet['BQ'+ str(row)] = _cell_value(day_json.get("V_tstn_lodochny"))
    excel_sheet['BQ'+ str(row)].number_format = "0.000"
    match _status(day_json.get("V_tstn_lodochny")):
        case 0:
            excel_sheet['BQ'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BQ'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BQ'+ str(row)].fill = red_fill
    
    excel_sheet['BR'+ str(row)] = _cell_value(day_json.get("V_tstn_tagul"))
    excel_sheet['BR'+ str(row)].number_format = "0.00"
    match _status(day_json.get("V_tstn_tagul")):
        case 0:
            excel_sheet['BR'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BR'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BR'+ str(row)].fill = red_fill
    
    excel_sheet['BS'+ str(row)] = _cell_value(day_json.get("V_tstn_skn"))
    match _status(day_json.get("V_tstn_skn")):
        case 0:
            excel_sheet['BS'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BS'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BS'+ str(row)].fill = red_fill
    
    excel_sheet['BT'+ str(row)] = _cell_value(day_json.get("V_tstn_vo"))
    match _status(day_json.get("V_tstn_vo")):
        case 0:
            excel_sheet['BT'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BT'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BT'+ str(row)].fill = red_fill
    
    excel_sheet['BU'+ str(row)] = _cell_value(day_json.get("V_tstn_tng"))
    match _status(day_json.get("V_tstn_tng")):
        case 0:
            excel_sheet['BU'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BU'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BU'+ str(row)].fill = red_fill
    
    excel_sheet['BV'+ str(row)] = _cell_value(day_json.get("V_tstn_kchng"))
    match _status(day_json.get("V_tstn_kchng")):
        case 0:
            excel_sheet['BV'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['BV'+ str(row)].fill = green_fill
        case _:
            excel_sheet['BV'+ str(row)].fill = red_fill

    if row == 4:
        for col in ["BW", "BX", "BY", "BZ", "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH"]:
            cell = excel_sheet[f"{col}{row}"]
            cell.value = None
            cell.fill = pale_green_fill
    else:
        excel_sheet['BW'+ str(row)] = _cell_value(day_json.get("F"))
        match _status(day_json.get("F")):
            case 0:
                excel_sheet['BW'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['BW'+ str(row)].fill = green_fill
            case _:
                excel_sheet['BW'+ str(row)].fill = red_fill
        
        excel_sheet['BX'+ str(row)] = _cell_value(day_json.get("F_vn"))
        match _status(day_json.get("F_vn")):
            case 0:
                excel_sheet['BX'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['BX'+ str(row)].fill = green_fill
            case _:
                excel_sheet['BX'+ str(row)].fill = red_fill
        
        excel_sheet['BY'+ str(row)] = _cell_value(day_json.get("F_suzun"))
        match _status(day_json.get("F_suzun")):
            case 0:
                excel_sheet['BY'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['BY'+ str(row)].fill = green_fill
            case _:
                excel_sheet['BY'+ str(row)].fill = red_fill
        
        excel_sheet['BZ'+ str(row)] = _cell_value(day_json.get("F_suzun_vankor"))
        match _status(day_json.get("F_suzun_vankor")):
            case 0:
                excel_sheet['BZ'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['BZ'+ str(row)].fill = green_fill
            case _:
                excel_sheet['BZ'+ str(row)].fill = red_fill
        
        excel_sheet['CA'+ str(row)] = _cell_value(day_json.get("F_suzun_vslu"))
        match _status(day_json.get("F_suzun_vslu")):
            case 0:
                excel_sheet['CA'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['CA'+ str(row)].fill = green_fill
            case _:
                excel_sheet['CA'+ str(row)].fill = red_fill
        
        excel_sheet['CB'+ str(row)] = _cell_value(day_json.get("F_tagul"))
        match _status(day_json.get("F_tagul")):
            case 0:
                excel_sheet['CB'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['CB'+ str(row)].fill = green_fill
            case _:
                excel_sheet['CB'+ str(row)].fill = red_fill
        
        excel_sheet['CC'+ str(row)] = _cell_value(day_json.get("F_tagul_lpu"))
        match _status(day_json.get("F_tagul_lpu")):
            case 0:
                excel_sheet['CC'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['CC'+ str(row)].fill = green_fill
            case _:
                excel_sheet['CC'+ str(row)].fill = red_fill
        
        excel_sheet['CD'+ str(row)] = _cell_value(day_json.get("F_tagul_tpu"))
        match _status(day_json.get("F_tagul_tpu")):
            case 0:
                excel_sheet['CD'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['CD'+ str(row)].fill = green_fill
            case _:
                excel_sheet['CD'+ str(row)].fill = red_fill
        
        excel_sheet['CE'+ str(row)] = _cell_value(day_json.get("F_skn"))
        match _status(day_json.get("F_skn")):
            case 0:
                excel_sheet['CE'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['CE'+ str(row)].fill = green_fill
            case _:
                excel_sheet['CE'+ str(row)].fill = red_fill
        
        excel_sheet['CF'+ str(row)] = _cell_value(day_json.get("F_vo"))
        match _status(day_json.get("F_vo")):
            case 0:
                excel_sheet['CF'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['CF'+ str(row)].fill = green_fill
            case _:
                excel_sheet['CF'+ str(row)].fill = red_fill
        
        excel_sheet['CG'+ str(row)] = _cell_value(day_json.get("F_tng"))
        match _status(day_json.get("F_tng")):
            case 0:
                excel_sheet['CG'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['CG'+ str(row)].fill = green_fill
            case _:
                excel_sheet['CG'+ str(row)].fill = red_fill
        
        excel_sheet['CH'+ str(row)] = _cell_value(day_json.get("F_kchng"))
        match _status(day_json.get("F_kchng")):
            case 0:
                excel_sheet['CH'+ str(row)].fill = pale_green_fill
            case 1:
                excel_sheet['CH'+ str(row)].fill = green_fill
            case _:
                excel_sheet['CH'+ str(row)].fill = red_fill

    excel_sheet['CJ'+ str(row)] = _cell_value(day_json.get("Q_gnps"))
    match _status(day_json.get("Q_gnps")):
        case 0:
            excel_sheet['CJ'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['CJ'+ str(row)].fill = green_fill
        case _:
            excel_sheet['CJ'+ str(row)].fill = red_fill
    
    excel_sheet['CK'+ str(row)] = _cell_value(day_json.get("Q_nps_1_2"))
    match _status(day_json.get("Q_nps_1_2")):
        case 0:
            excel_sheet['CK'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['CK'+ str(row)].fill = green_fill
        case _:
            excel_sheet['CK'+ str(row)].fill = red_fill
    
    excel_sheet['CL'+ str(row)] = _cell_value(day_json.get("Q_knps"))
    match _status(day_json.get("Q_knps")):
        case 0:
            excel_sheet['CL'+ str(row)].fill = pale_green_fill
        case 1:
            excel_sheet['CL'+ str(row)].fill = green_fill
        case _:
            excel_sheet['CL'+ str(row)].fill = red_fill
    
    row += 1


# Записываем в Excel значения по окончании месяца
excel_sheet['A'+ str(row)] = 'накоп'
excel_sheet['A'+ str(row)].fill = pale_blue_fill

excel_sheet['AI'+ str(row)] = _m("G_per_month")
excel_sheet['AI'+ str(row)].fill = pale_green_fill

excel_sheet['D'+ str(row)] = _m("Q_vankor_month")
excel_sheet['D'+ str(row)].fill = pale_green_fill

excel_sheet['T'+ str(row)] = _m("Q_suzun_month")
excel_sheet['T'+ str(row)].fill = pale_green_fill

excel_sheet['U'+ str(row)] = _m("Q_vslu_month")
excel_sheet['U'+ str(row)].fill = pale_green_fill

excel_sheet['AE'+ str(row)] = _m("Q_tng_month")
excel_sheet['AE'+ str(row)].fill = pale_green_fill

excel_sheet['AY'+ str(row)] = _m("Q_vo_month")
excel_sheet['AY'+ str(row)].fill = pale_green_fill

excel_sheet['AB'+ str(row)] = _m("G_suzun_vslu_month")
excel_sheet['AB'+ str(row)].fill = pale_green_fill

excel_sheet['AA'+ str(row)] = _m("G_suzun_slu_month")
excel_sheet['AA'+ str(row)].fill = pale_green_fill

excel_sheet['Z'+ str(row)] = _m("G_suzun_month")
excel_sheet['Z'+ str(row)].fill = pale_green_fill

excel_sheet['AD'+ str(row)] = _m("delta_G_suzun_month")
excel_sheet['AD'+ str(row)].fill = pale_green_fill

excel_sheet['AZ'+ str(row)] = _m("G_upn_lodochny_ichem_month")
excel_sheet['AZ'+ str(row)].fill = pale_green_fill

excel_sheet['BB'+ str(row)] = _m("Q_kchng_month")
excel_sheet['BB'+ str(row)].fill = pale_green_fill

excel_sheet['BC'+ str(row)] = _m("G_kchng_month")
excel_sheet['BC'+ str(row)].fill = pale_green_fill

excel_sheet['AK'+ str(row)] = _m("Q_tagul_month")
excel_sheet['AK'+ str(row)].fill = pale_green_fill

excel_sheet['AP'+ str(row)] = _m("Q_lodochny_month")
excel_sheet['AP'+ str(row)].fill = pale_green_fill

g_lodochny_upsv_yu_month = _m("G_lodochny_upsv_yu_month")
if g_lodochny_upsv_yu_month is None:
    g_lodochny_upsv_yu_month = _m("G_lodochny_uspv_yu_month")
excel_sheet['AQ'+ str(row)] = g_lodochny_upsv_yu_month
excel_sheet['AQ'+ str(row)].fill = pale_green_fill

excel_sheet['AU'+ str(row)] = _m("G_upn_lodochny_month")
excel_sheet['AU'+ str(row)].fill = pale_green_fill

excel_sheet['AX'+ str(row)] = _m("delta_G_upn_lodochny_month")
excel_sheet['AX'+ str(row)].fill = pale_green_fill

excel_sheet['AM'+ str(row)] = _m("G_tagul_month")
excel_sheet['AM'+ str(row)].fill = pale_green_fill

excel_sheet['AN'+ str(row)] = _m("delta_G_tagul_month")
excel_sheet['AN'+ str(row)].fill = pale_green_fill

excel_sheet['BA'+ str(row)] = _m("G_tagul_lodochny_month")
excel_sheet['BA'+ str(row)].fill = pale_green_fill

excel_sheet['AV'+ str(row)] = _m("G_lodochny_month")
excel_sheet['AV'+ str(row)].fill = pale_green_fill

excel_sheet['O'+ str(row)] = _m("G_sikn_vslu_month")
excel_sheet['O'+ str(row)].fill = pale_green_fill

excel_sheet['P'+ str(row)] = _m("G_sikn_tagul_month")
excel_sheet['P'+ str(row)].fill = pale_green_fill

excel_sheet['N'+ str(row)] = _m("G_sikn_suzun_month")
excel_sheet['N'+ str(row)].fill = pale_green_fill

excel_sheet['Q'+ str(row)] = _m("G_sikn_tng_month")
excel_sheet['Q'+ str(row)].fill = pale_green_fill

excel_sheet['L'+ str(row)] = _m("G_sikn_month")
excel_sheet['L'+ str(row)].fill = pale_green_fill

excel_sheet['M'+ str(row)] = _m("G_sikn_vankor_month")
excel_sheet['M'+ str(row)].fill = pale_green_fill

excel_sheet['BD'+ str(row)] = _m("G_skn_month")
excel_sheet['BD'+ str(row)].fill = pale_green_fill

excel_sheet['R'+ str(row)] = _m("delta_G_sikn_month")
excel_sheet['R'+ str(row)].fill = pale_green_fill

excel_sheet['BZ'+ str(row)] = _m("F_suzun_vankor_sum")
excel_sheet['BZ'+ str(row)].fill = pale_green_fill

excel_sheet['CA'+ str(row)] = _m("F_suzun_vslu_sum")
excel_sheet['CA'+ str(row)].fill = pale_green_fill

excel_sheet['CC'+ str(row)] = _m("F_tagul_lpu_sum")
excel_sheet['CC'+ str(row)].fill = pale_green_fill

excel_sheet['CD'+ str(row)] = _m("F_tagul_tpu_sum")
excel_sheet['CD'+ str(row)].fill = pale_green_fill

excel_sheet['CB'+ str(row)] = _m("F_tagul_sum")
excel_sheet['CB'+ str(row)].fill = pale_green_fill

excel_sheet['CE'+ str(row)] = _m("F_skn_sum")
excel_sheet['CE'+ str(row)].fill = pale_green_fill

excel_sheet['CF'+ str(row)] = _m("F_vo_sum")
excel_sheet['CF'+ str(row)].fill = pale_green_fill

excel_sheet['CG'+ str(row)] = _m("F_tng_sum")
excel_sheet['CG'+ str(row)].fill = pale_green_fill

excel_sheet['CH'+ str(row)] = _m("F_kchng_sum")
excel_sheet['CH'+ str(row)].fill = pale_green_fill

excel_sheet['BW'+ str(row)] = _m("F_sum")
excel_sheet['BW'+ str(row)].fill = pale_green_fill

excel_sheet['BE'+ str(row)] = _m("G_gnps_month")
excel_sheet['BE'+ str(row)].fill = pale_green_fill

excel_sheet['AG'+ str(row)] = _m("G_buying_oil_month")
excel_sheet['AG'+ str(row)].fill = pale_blue_fill

excel_sheet['AH'+ str(row)] = _m("G_out_udt_month")
excel_sheet['AH'+ str(row)].fill = pale_blue_fill

excel_sheet['BY'+ str(row)] = _m("F_suzun_sum")
excel_sheet['BY'+ str(row)].fill = pale_green_fill

excel_sheet['BX'+ str(row)] = _m("F_vn_sum")
excel_sheet['BX'+ str(row)].fill = pale_green_fill


# Сохраняем результат
suffix = "_auto" if start_autobalance else ""
base_name = f"balance_{date_str}{suffix}.xlsx"
try:
    excel_file.save(base_name)
except PermissionError:
    from pathlib import Path

    base_path = Path(base_name)
    saved = False
    for i in range(1, 100):
        alt_path = base_path.with_name(f"{base_path.stem}_copy{i}{base_path.suffix}")
        try:
            excel_file.save(str(alt_path))
            saved = True
            break
        except PermissionError:
            continue
    if not saved:
        raise


